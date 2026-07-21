"""Build empirical prop-outcome dataset from tennis-data.co.uk workbooks.

Each row of the output = one completed match with:
  tour, date, surface, best_of, level (tourney tier),
  p_win  (devigged Pinnacle closing prob of the eventual winner),
  p_fav  (devigged prob of the pre-match favorite),
  fav_won, set score, set1 winner (favorite?), total games,
  game margin (favorite - dog), any tiebreak, retired flag.

Used to compute empirical conditional prop distributions given the match
line — the model-free benchmark for Kalshi prop prices.
"""
import glob
import json
import os
import sys
import openpyxl

SNAP = sys.argv[1] if len(sys.argv) > 1 else "."


def rows_from(path, tour):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(it)]
    ix = {h: i for i, h in enumerate(header)}

    def g(row, col):
        i = ix.get(col)
        return row[i] if i is not None and i < len(row) else None

    out = []
    for row in it:
        try:
            psw, psl = g(row, "PSW"), g(row, "PSL")
            if not psw or not psl:
                psw, psl = g(row, "AvgW"), g(row, "AvgL")
            if not psw or not psl:
                continue
            psw, psl = float(psw), float(psl)
            comment = str(g(row, "Comment") or "").strip().lower()
            best_of = int(g(row, "Best of") or 3)
            wsets, lsets = g(row, "Wsets"), g(row, "Lsets")
            sets = []
            games_w = games_l = 0
            tb = False
            for k in range(1, 6):
                a, b = g(row, f"W{k}"), g(row, f"L{k}")
                if a is None or b is None:
                    break
                a, b = int(a), int(b)
                sets.append((a, b))
                games_w += a
                games_l += b
                if (a, b) in ((7, 6), (6, 7)):
                    tb = True
            # devig winner prob (multiplicative)
            iw, il = 1 / psw, 1 / psl
            p_win = iw / (iw + il)
            fav_is_winner = p_win >= 0.5
            p_fav = p_win if fav_is_winner else 1 - p_win
            set1 = None
            if sets:
                set1 = (sets[0][0] > sets[0][1]) == fav_is_winner
            rec = {
                "tour": tour,
                "date": str(g(row, "Date"))[:10],
                "surface": str(g(row, "Surface") or ""),
                "series": str(g(row, "Series") or g(row, "Tier") or ""),
                "round": str(g(row, "Round") or ""),
                "best_of": best_of,
                "p_win": round(p_win, 4),
                "p_fav": round(p_fav, 4),
                "fav_won": fav_is_winner,
                "wsets": int(wsets) if wsets is not None else None,
                "lsets": int(lsets) if lsets is not None else None,
                "sets": sets,
                "total_games": games_w + games_l,
                "margin_winner": games_w - games_l,
                "any_tb": tb,
                "retired": "retired" in comment or "walkover" in comment,
                "completed": comment in ("completed", "compleed", ""),
            }
            out.append(rec)
        except Exception:
            continue
    return out


all_rows = []
for path in sorted(glob.glob(os.path.join(SNAP, "tennisdata", "*.xlsx"))):
    tour = "ATP" if os.path.basename(path).startswith("atp") else "WTA"
    r = rows_from(path, tour)
    all_rows += r
    print(os.path.basename(path), len(r))

json.dump(all_rows, open(os.path.join(SNAP, "empirical_matches.json"), "w"))
print("total", len(all_rows))
